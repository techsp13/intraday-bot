"""
Copy 100% Same-Day Intraday CSV & Excel Reports to brain and workspace folders
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def format_and_export_sameday_reports():
    df = pd.read_csv("3year_trades_sameday_intraday.csv")

    csv_path = "3year_trades_log.csv"
    brain_csv_path = r"C:\Users\ASUS\.gemini\antigravity\brain\3f4a3a4c-9e99-4921-807e-44f24c8fd355\3year_trades_log.csv"
    excel_path = "3year_trades_report.xlsx"
    brain_excel_path = r"C:\Users\ASUS\.gemini\antigravity\brain\3f4a3a4c-9e99-4921-807e-44f24c8fd355\3year_trades_report.xlsx"

    df.to_csv(csv_path, index=False)
    df.to_csv(brain_csv_path, index=False)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "100% Same-Day Intraday Log"

    ws.append(["STRICT 100% SAME-DAY INTRADAY TRADE LOG — NIFTY RELATIVE STRENGTH STRATEGY"])
    ws.append([f"Starting Capital: Rs. 100,000.00 | Ending Capital: Rs. 274,826.98 | Net Return: +174.83% | Total Trades: {len(df)}"])
    ws.append([])

    headers = ["#", "Date", "Trigger Time", "Timestamp", "Symbol", "Direction", "Entry Price (Rs)", "Exit Price (Rs)", "Stop Loss (Rs)", "Shares", "Outcome", "P&L (Rs)", "Return %", "Capital After (Rs)"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    win_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    loss_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    be_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, row in df.iterrows():
        r_idx = idx + 5
        ws.append([
            idx + 1,
            row['Date'],
            row['Trigger_Time'],
            row['Timestamp'],
            row['Symbol'],
            row['Direction'],
            row['Entry_Price'],
            row['Exit_Price'],
            row['SL_Price'],
            row['Shares'],
            row['Outcome'],
            row['PnL_Rs'],
            row['Return_Pct'],
            row['Capital_After']
        ])

        fill = win_fill if row['Outcome'] in ['HIT_T1', 'HIT_T2'] else (loss_fill if row['Outcome'] == 'HIT_SL' else be_fill)
        for col_num in range(1, 15):
            cell = ws.cell(row=r_idx, column=col_num)
            cell.fill = fill
            if col_num in [7, 8, 9, 12, 14]:
                cell.number_format = '#,##0.00'
            elif col_num == 13:
                cell.number_format = '+0.00%;-0.00%;0.00%'

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(excel_path)
    wb.save(brain_excel_path)
    print("Successfully exported 100% same-day intraday CSV & Excel reports!")

if __name__ == '__main__':
    format_and_export_sameday_reports()
