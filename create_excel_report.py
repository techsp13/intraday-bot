"""
Update create_excel_report.py with corrected Return % formatting
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def generate_formatted_excel():
    df = pd.read_csv("1year_trades_log.csv")

    excel_path = "1year_trades_report.xlsx"
    brain_excel_path = r"C:\Users\ASUS\.gemini\antigravity\brain\3f4a3a4c-9e99-4921-807e-44f24c8fd355\1year_trades_report.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1-Year Trade Log"

    ws.append(["1-YEAR HISTORICAL TRADE LOG — NIFTY RELATIVE STRENGTH STRATEGY"])
    ws.append(["Starting Capital: Rs. 100,000.00 | Ending Capital: Rs. 119,972.71 | Net Return: +19.97% | Total Trades: 392"])
    ws.append([])

    headers = ["#", "Date", "Symbol", "Direction", "Entry Price (Rs)", "Exit Price (Rs)", "Stop Loss (Rs)", "Shares", "Outcome", "P&L (Rs)", "Return %", "Capital After (Rs)"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    win_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    loss_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    be_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, row in df.iterrows():
        r_idx = idx + 5
        ws.append([
            idx + 1,
            row['Date'],
            row['Symbol'],
            row['Direction'],
            row['Entry_Price'],
            row['Exit_Price'],
            row['SL_Price'],
            row['Shares'],
            row['Outcome'],
            row['PnL_Rs'],
            row['Return_Pct'],  # Stored as decimal e.g. 0.015
            row['Capital_After']
        ])

        fill = win_fill if row['Outcome'] in ['HIT_T1', 'HIT_T2'] else (loss_fill if row['Outcome'] == 'HIT_SL' else be_fill)
        for col_num in range(1, 13):
            cell = ws.cell(row=r_idx, column=col_num)
            cell.fill = fill
            if col_num in [5, 6, 7, 10, 12]:
                cell.number_format = '#,##0.00'
            elif col_num == 11:
                cell.number_format = '+0.00%;-0.00%;0.00%'

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(excel_path)
    wb.save(brain_excel_path)
    print(f"Excel file updated at {excel_path} and {brain_excel_path}")

if __name__ == '__main__':
    generate_formatted_excel()
