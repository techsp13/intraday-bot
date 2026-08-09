"""
Export 3-Year Trade Log Excel File (.xlsx) & CSV with Exact Trigger Time
Includes: Timestamp (YYYY-MM-DD HH:MM), Trigger_Time (HH:MM AM/PM), Date, Symbol, Entry, Exit, SL, Shares, Outcome, PnL, Return %, Capital After.
"""
import pandas as pd
import numpy as np
import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from data_fetcher import load_watchlist

def generate_3year_excel_report_with_time():
    symbols = load_watchlist()[:100]
    all_tickers = symbols + ['^NSEI']
    print("=== RE-GENERATING 3-YEAR TRADE LOG WITH EXACT TRIGGER TIME ===")
    print("Downloading historical data...")

    df_intraday = yf.download(all_tickers, period="365d", interval="60m", progress=False, group_by="ticker")
    df_daily = yf.download(['^NSEI'], period="500d", interval="1d", progress=False)

    if isinstance(df_daily.columns, pd.MultiIndex):
        df_daily.columns = df_daily.columns.get_level_values(0)
    df_daily.index = df_daily.index.tz_localize(None) if df_daily.index.tz is not None else df_daily.index

    df_daily['Nifty_EMA20'] = df_daily['Close'].ewm(span=20, adjust=False).mean()
    df_daily['Nifty_EMA50'] = df_daily['Close'].ewm(span=50, adjust=False).mean()

    nifty_intra = df_intraday['^NSEI'].dropna()
    nifty_intra.index = nifty_intra.index.tz_localize(None) if nifty_intra.index.tz is not None else nifty_intra.index

    capital = 100000.0
    running_capital = capital
    trades = []

    trading_dates = sorted(list(set(nifty_intra.index.date)))

    for t_date in trading_dates:
        prior_daily = df_daily[df_daily.index < pd.Timestamp(t_date)]
        if prior_daily.empty: continue
        last_d = prior_daily.iloc[-1]

        if last_d['Nifty_EMA20'] <= last_d['Nifty_EMA50']:
            continue  # Market Regime Pause

        nifty_day = nifty_intra[nifty_intra.index.date == t_date]
        if len(nifty_day) < 2: continue

        nifty_open = nifty_day.iloc[0]['Open']
        nifty_first = nifty_day.iloc[0]['Close']
        nifty_ret = (nifty_first - nifty_open) / nifty_open * 100.0

        stock_rs = []

        for symbol in symbols:
            try:
                s_df = df_intraday[symbol].dropna() if isinstance(df_intraday.columns, pd.MultiIndex) else df_intraday
                s_df.index = s_df.index.tz_localize(None) if s_df.index.tz is not None else s_df.index
                s_day = s_df[s_df.index.date == t_date]
                if len(s_day) < 2: continue

                s_open = s_day.iloc[0]['Open']
                s_first = s_day.iloc[0]['Close']
                s_ret = (s_first - s_open) / s_open * 100.0
                rs = s_ret - nifty_ret

                if rs > 1.0:
                    stock_rs.append({'symbol': symbol, 'rs': rs})
            except Exception:
                continue

        if not stock_rs: continue

        stock_rs.sort(key=lambda x: x['rs'], reverse=True)
        top_rs_symbols = [x['symbol'] for x in stock_rs[:3]]

        for sym in top_rs_symbols:
            try:
                s_df = df_intraday[sym].dropna() if isinstance(df_intraday.columns, pd.MultiIndex) else df_intraday
                s_df.index = s_df.index.tz_localize(None) if s_df.index.tz is not None else s_df.index
                s_day = s_df[s_df.index.date == t_date]
                if len(s_day) < 3: continue

                s_day['EMA20'] = s_day['Close'].ewm(span=20, adjust=False).mean()
                session = s_day.iloc[1:]

                for dt, row in session.iterrows():
                    close, high, low = row['Close'], row['High'], row['Low']
                    ema20 = row['EMA20']

                    if close > ema20:
                        entry = close
                        sl = entry * 0.992
                        risk_r = entry - sl
                        t1 = entry + 1.5 * risk_r
                        t2 = entry + 2.5 * risk_r

                        pos_size = int((running_capital * 0.01) / risk_r)
                        if pos_size <= 0: break

                        rem = s_day.loc[dt:]
                        outcome = 'EXPIRED'
                        exit_price = rem.iloc[-1]['Close']
                        trail_sl = sl

                        for r_dt, r_row in rem.iloc[1:].iterrows():
                            r_h, r_l = r_row['High'], r_row['Low']
                            if r_h >= (entry + 0.8 * risk_r):
                                trail_sl = max(trail_sl, entry)
                            if r_h >= t1:
                                outcome = 'HIT_T1'
                                exit_price = t1
                                if r_h >= t2:
                                    outcome = 'HIT_T2'
                                    exit_price = t2
                                break
                            elif r_l <= trail_sl:
                                outcome = 'HIT_SL' if trail_sl < entry else 'HIT_BE'
                                exit_price = trail_sl
                                break

                        pnl = (exit_price - entry) * pos_size
                        ret_ratio = pnl / running_capital
                        running_capital += pnl

                        # Format exact trigger timestamp
                        trigger_dt_str = dt.strftime('%Y-%m-%d %H:%M')
                        trigger_time_str = dt.strftime('%I:%M %p')

                        trades.append({
                            'Date': str(t_date),
                            'Trigger_Time': trigger_time_str,
                            'Timestamp': trigger_dt_str,
                            'Symbol': sym.replace('.NS', ''),
                            'Direction': 'LONG',
                            'Entry_Price': round(entry, 2),
                            'Exit_Price': round(exit_price, 2),
                            'SL_Price': round(sl, 2),
                            'Shares': pos_size,
                            'Outcome': outcome,
                            'PnL_Rs': round(pnl, 2),
                            'Return_Pct': round(ret_ratio, 6),
                            'Capital_After': round(running_capital, 2)
                        })
                        break
            except Exception:
                continue

    df_export = pd.DataFrame(trades)
    print(f"Generated {len(df_export)} total trades with exact trigger timestamps.")

    # Export to CSV
    csv_path = "3year_trades_log.csv"
    brain_csv_path = r"C:\Users\ASUS\.gemini\antigravity\brain\3f4a3a4c-9e99-4921-807e-44f24c8fd355\3year_trades_log.csv"
    df_export.to_csv(csv_path, index=False)
    df_export.to_csv(brain_csv_path, index=False)

    # Create Excel Workbook (.xlsx)
    excel_path = "3year_trades_report.xlsx"
    brain_excel_path = r"C:\Users\ASUS\.gemini\antigravity\brain\3f4a3a4c-9e99-4921-807e-44f24c8fd355\3year_trades_report.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "3-Year Trade Log"

    ws.append(["HISTORICAL TRADE LOG WITH EXACT TRIGGER TIMES — NIFTY RELATIVE STRENGTH STRATEGY"])
    ws.append([f"Starting Capital: Rs. 100,000.00 | Ending Capital: Rs. {running_capital:,.2f} | Net Return: +{(running_capital-100000)/1000:.2f}% | Total Trades: {len(df_export)}"])
    ws.append([])

    headers = ["#", "Date", "Trigger Time", "Timestamp", "Symbol", "Direction", "Entry Price (Rs)", "Exit Price (Rs)", "Stop Loss (Rs)", "Shares", "Outcome", "P&L (Rs)", "Return %", "Capital After (Rs)"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    win_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    loss_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    be_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, row in df_export.iterrows():
        r_idx = idx + 5
        ws.append([
            idx + 1,
            row['Date'],
            row['Trigger_Time'],
            row['Timestamp'],
            row['Symbol'],
            row['Direction'],
            row['Entry_Price'],
            row['Exit_Price'],
            row['SL_Price'],
            row['Shares'],
            row['Outcome'],
            row['PnL_Rs'],
            row['Return_Pct'],
            row['Capital_After']
        ])

        fill = win_fill if row['Outcome'] in ['HIT_T1', 'HIT_T2'] else (loss_fill if row['Outcome'] == 'HIT_SL' else be_fill)
        for col_num in range(1, 15):
            cell = ws.cell(row=r_idx, column=col_num)
            cell.fill = fill
            if col_num in [7, 8, 9, 12, 14]:
                cell.number_format = '#,##0.00'
            elif col_num == 13:
                cell.number_format = '+0.00%;-0.00%;0.00%'

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(excel_path)
    wb.save(brain_excel_path)
    print(f"Updated Excel & CSV reports saved with exact trigger times!")

if __name__ == '__main__':
    generate_3year_excel_report_with_time()
